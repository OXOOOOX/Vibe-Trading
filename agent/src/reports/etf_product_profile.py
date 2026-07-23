"""Traceable ETF product, index-methodology, share and peer-flow snapshots.

The service deliberately stores observations as immutable ETF research snapshots.
Reports bind the composite profile returned here, so a later refresh cannot rewrite
an older report's product facts.  Dynamic flow estimates are labelled as estimates;
exchange-published fund units remain the primary observation.
"""

from __future__ import annotations

import hashlib
import io
import json
import re
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from html import unescape
from typing import Any, Callable, Iterable

import requests

from src.research.source_ingestion import CollectedSource, SourceIngestionService

from .contracts import ETFResearchSnapshot, utc_now
from .etf_research import ETFResearchStore, build_etf_snapshot, stable_fingerprint
from .etf_source_registry import (
    ETFSourceRegistry,
    ETFSourceRule,
    get_etf_source_registry,
    source_context,
)
from .etf_universe_provider import (
    AuditedETFIndexMapper,
    _qualify_index_code,
    normalize_etf_symbol,
)


_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/124.0 Safari/537.36"
    )
}

_AUDITED_588870 = {
    "identity": {
        "fund_full_name": "汇添富上证科创板50成份交易型开放式指数证券投资基金",
        "fund_short_name": "汇添富上证科创板50成份ETF",
        "exchange_short_name": "科创50指",
        "manager": "汇添富基金管理股份有限公司",
        "custodian": "中信证券股份有限公司",
        "exchange": "上海证券交易所",
        "contract_effective_date": "2025-01-20",
        "listing_date": "2025-01-27",
        "tracked_index_code": "000688.SH",
        "tracked_index_name": "上证科创板50成份指数",
    },
    "index_methodology": {
        "index_code": "000688.SH",
        "index_name": "上证科创板50成份指数",
        "version": "V1.1",
        "published_at": "2020-12",
        "target_component_count": 50,
        "single_constituent_weight_cap": 0.10,
        "top_five_weight_cap": 0.40,
        "review_frequency": "quarterly",
        "review_months": [3, 6, 9, 12],
        "regular_rebalance_change_cap": 0.10,
        "buffer_entry_rank": 40,
        "buffer_retention_rank": 60,
    },
    "reported_metrics": {
        "management_fee_rate": 0.0015,
        "custody_fee_rate": 0.0005,
        "published_net_assets": 710_743_392.87,
        "published_fund_units": 518_552_000.0,
        "published_unit_nav": 1.3706,
    },
    "reported_as_of": "2025-12-31",
}

_INDEX_NAME_ALIASES = {
    "000688.SH": ("科创50", "科创板50"),
    "000300.SH": ("沪深300",),
    "000852.SH": ("中证1000",),
    "000016.SH": ("上证50",),
    "931787.CSI": ("中证香港创新药", "香港创新药", "港股创新药", "港股通创新药", "HK创新药"),
}

_SZSE_INDEX_CODE_FIELDS = (
    "标的指数代码",
    "跟踪指数代码",
    "跟踪标的指数代码",
    "指数代码",
)
_SZSE_INDEX_NAME_FIELDS = (
    "标的指数名称",
    "跟踪指数名称",
    "跟踪标的指数",
    "标的指数",
    "跟踪指数",
    "指数名称",
)
_SZSE_FUND_CODE_FIELDS = ("基金代码", "证券代码", "交易代码")
_SZSE_FUND_NAME_FIELDS = ("基金简称", "基金扩位简称", "证券简称")
_SZSE_FUND_UNITS_FIELDS = ("当前规模(份)", "基金份额(份)", "基金份额", "份额")
_SZSE_MARKET_PRICE_FIELDS = ("最新价", "收盘价", "市场价格", "现价")
_SZSE_NAV_FIELDS = ("净值", "基金份额净值", "单位净值")


def _row_value(row: dict[str, Any], fields: Iterable[str]) -> Any:
    for field in fields:
        value = row.get(field)
        if value is not None and str(value).strip():
            return value
    return None


def _compact_index_code(value: Any) -> str:
    return re.sub(r"[^0-9A-Z]", "", str(value or "").upper().split(".", 1)[0])


def _compact_index_name(value: Any) -> str:
    return re.sub(r"[^0-9A-Z\u4e00-\u9fff]", "", str(value or "").upper())


def _index_name_aliases(index_code: str, index_name: str) -> tuple[str, ...]:
    """Build reusable cross-exchange aliases without making peer selection symbol-specific."""

    tracked_code = _compact_index_code(index_code)
    aliases = {
        _compact_index_name(item)
        for configured_code, configured_aliases in _INDEX_NAME_ALIASES.items()
        if _compact_index_code(configured_code) == tracked_code
        for item in configured_aliases
        if _compact_index_name(item)
    }
    canonical = _compact_index_name(index_name)
    if canonical:
        aliases.add(canonical)
        base = re.sub(r"(?:全收益)?指数$", "", canonical)
        base = re.sub(r"成份$", "", base)
        if base:
            aliases.add(base)
            without_provider = re.sub(r"^(?:中证|上证|深证)", "", base)
            if len(without_provider) >= 4:
                aliases.add(without_provider)
            if "香港" in base:
                aliases.add(base.replace("香港", "港股"))
    # Very short aliases (for example just “50”) create false cross-index groups.
    return tuple(sorted((item for item in aliases if len(item) >= 4), key=lambda item: (-len(item), item)))


def _same_index_code(candidate: Any, tracked_index_code: str) -> bool:
    candidate_code = _compact_index_code(candidate)
    tracked_code = _compact_index_code(tracked_index_code)
    if not candidate_code or not tracked_code:
        return False
    if candidate_code == tracked_code:
        return True
    # 沪深300 uses 000300 on SSE/CSI disclosures and may appear as 399300 on SZSE.
    return {candidate_code, tracked_code} == {"000300", "399300"}


def _same_index_name(candidate: Any, aliases: Iterable[str]) -> bool:
    compact = _compact_index_name(candidate)
    if not compact:
        return False
    return any(compact == alias for alias in aliases)


def _fund_name_matches_index(candidate: Any, aliases: Iterable[str]) -> bool:
    compact = _compact_index_name(candidate)
    if not compact:
        return False
    return any(alias in compact for alias in aliases)

_FIELD_LABELS = {
    "fund_full_name": "基金全称",
    "fund_short_name": "基金简称",
    "manager": "基金管理人",
    "custodian": "基金托管人",
    "exchange": "上市交易所",
    "contract_effective_date": "合同生效日",
    "listing_date": "上市日",
    "tracked_index_code": "跟踪指数代码",
    "tracked_index_name": "跟踪指数名称",
    "index_code": "指数代码",
    "index_name": "指数名称",
    "version": "规则版本",
    "published_at": "规则发布日期",
    "target_component_count": "目标成分数量",
    "single_constituent_weight_cap": "单一成分权重上限",
    "top_five_weight_cap": "前五大成分权重上限",
    "review_frequency": "定期调样频率",
    "management_fee_rate": "管理费率",
    "custody_fee_rate": "托管费率",
    "unit_nav": "单位净值",
    "cumulative_nav": "累计净值",
    "fund_units": "基金份额",
    "published_net_assets": "已披露基金资产净值",
    "exchange_market_value": "场内市值（价格×上市份额）",
    "iopv": "IOPV",
    "premium_discount_rate": "折溢价率",
}

_FIELD_UNITS = {
    "management_fee_rate": "ratio",
    "custody_fee_rate": "ratio",
    "unit_nav": "CNY_per_fund_unit",
    "cumulative_nav": "CNY_per_fund_unit",
    "fund_units": "fund_units",
    "published_fund_units": "fund_units",
    "published_net_assets": "CNY",
    "exchange_market_value": "CNY",
    "iopv": "CNY_per_fund_unit",
    "premium_discount_rate": "ratio",
    "single_constituent_weight_cap": "ratio",
    "top_five_weight_cap": "ratio",
    "target_component_count": "count",
}


def _hash_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _hash_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    text = str(value).strip().replace(",", "").replace("%", "")
    if not text or text in {"--", "-", "nan", "None"}:
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _iso_date(value: Any) -> str | None:
    text = str(value or "").strip()
    match = re.search(r"(20\d{2})[-年 /](\d{1,2})[-月 /](\d{1,2})", text)
    if not match:
        return None
    return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"


def _as_of_date(value: str | None) -> date:
    text = str(value or "").strip()
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            return datetime.now(timezone.utc).date()


def decode_official_html(raw: bytes, declared_encoding: str | None = None) -> tuple[str, str]:
    """Decode manager pages that may declare GBK while serving GB18030 bytes."""

    candidates = [declared_encoding, "utf-8", "gb18030", "gbk"]
    seen: set[str] = set()
    for candidate in candidates:
        encoding = str(candidate or "").strip().lower()
        if not encoding or encoding in seen:
            continue
        seen.add(encoding)
        try:
            return raw.decode(encoding), encoding
        except (LookupError, UnicodeDecodeError):
            continue
    return raw.decode("gb18030", errors="replace"), "gb18030-replace"


def extract_pdf_text(raw: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - packaging/runtime boundary
        raise RuntimeError("PDF text extraction failed: pypdf is not installed") from exc
    try:
        reader = PdfReader(io.BytesIO(raw))
        content = "\n".join(str(page.extract_text() or "") for page in reader.pages)
    except Exception as exc:  # pragma: no cover - dependency/runtime boundary
        raise RuntimeError(f"PDF text extraction failed with pypdf: {exc}") from exc
    if not content.strip():
        raise RuntimeError("PDF text extraction failed with pypdf: document has no extractable text")
    return content


def _plain_html(value: str) -> str:
    try:
        from bs4 import BeautifulSoup

        return "\n".join(
            line.strip()
            for line in BeautifulSoup(value, "html.parser").get_text("\n").splitlines()
            if line.strip()
        )
    except Exception:
        return re.sub(r"\s+", " ", unescape(re.sub(r"<[^>]+>", " ", value))).strip()


def _field(
    value: Any,
    *,
    source_ids: Iterable[str],
    data_as_of: str,
    semantics: str,
    unit: str | None = None,
    status: str | None = None,
    note: str | None = None,
    formula: str | None = None,
    input_metrics: Iterable[str] = (),
    calculation_version: str | None = None,
) -> dict[str, Any]:
    available = value is not None and str(value).strip() not in {"", "--"}
    return {
        "value": value if available else None,
        "status": status or ("available" if available else "missing"),
        "unit": unit,
        "data_as_of": data_as_of,
        "source_ids": sorted({str(item) for item in source_ids if str(item)}),
        "semantics": semantics,
        "note": note,
        "formula": formula,
        "input_metrics": [str(item) for item in input_metrics if str(item)],
        "calculation_version": calculation_version,
        "source_kind": "derived" if formula else "observed",
    }


def _metric_map(profile: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    return {
        str(item.get("key") or ""): dict(item)
        for item in ((profile or {}).get("metrics") or [])
        if isinstance(item, dict) and item.get("key")
    }


def _stable_source_catalog(sources: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    """Keep source identity in a snapshot without retrieval-time hash churn."""

    keys = (
        "source_id", "kind", "title", "publisher", "url", "content_hash",
        "published_at", "body_status", "verification_status", "metadata",
    )
    return [
        {key: item.get(key) for key in keys if item.get(key) is not None}
        for item in sources
        if isinstance(item, dict) and item.get("source_id")
    ]


def _section_as_of(section: dict[str, Any], fallback: str) -> str:
    values: list[str] = []
    for item in section.values():
        if not isinstance(item, dict):
            continue
        value = str(item.get("data_as_of") or "").strip()
        if re.fullmatch(r"20\d{2}-\d{2}", value):
            value += "-01"
        if value:
            values.append(value)
    return max(values) if values else fallback


def _finalize_product_profile_state(raw: dict[str, Any]) -> dict[str, Any]:
    """Recompute the public profile verdict after every source has settled."""

    identity = dict(raw.get("identity") or {})
    methodology = dict(raw.get("index_methodology") or {})
    metrics = dict(raw.get("product_metrics") or {})
    hard_keys = (
        (identity, "manager"),
        (identity, "exchange"),
        (identity, "tracked_index_code"),
        (identity, "tracked_index_name"),
        (methodology, "version"),
        (methodology, "source_url"),
    )
    optional_keys = (
        "management_fee_rate",
        "custody_fee_rate",
        "unit_nav",
        "fund_units",
        "published_net_assets",
        "exchange_market_value",
        "iopv",
        "premium_discount_rate",
    )
    raw["missing_hard_fields"] = [
        key
        for section, key in hard_keys
        if (section.get(key) or {}).get("status") != "available"
        or not (section.get(key) or {}).get("source_ids")
    ]
    raw["missing_optional_fields"] = [
        key
        for key in optional_keys
        if (metrics.get(key) or {}).get("status") != "available"
    ]
    raw["hard_gate_status"] = (
        "passed" if not raw["missing_hard_fields"] else "failed_validation"
    )
    raw["quality_status"] = (
        "failed_validation"
        if raw["missing_hard_fields"]
        else "passed_with_gaps"
        if raw["missing_optional_fields"] or raw.get("source_errors")
        else "passed"
    )
    raw["coverage_ratio"] = (
        1.0 - len(raw["missing_optional_fields"]) / len(optional_keys)
    )
    return raw


@dataclass(frozen=True)
class _FetchedSource:
    source_id: str
    kind: str
    title: str
    publisher: str
    url: str
    content: str
    content_hash: str
    retrieved_at: str
    published_at: str | None = None
    body_status: str = "full_text"
    verification_status: str = "official_primary"
    metadata: dict[str, Any] | None = None

    def public(self) -> dict[str, Any]:
        return {
            "source_id": self.source_id,
            "kind": self.kind,
            "title": self.title,
            "publisher": self.publisher,
            "url": self.url,
            "content_hash": self.content_hash,
            "retrieved_at": self.retrieved_at,
            "published_at": self.published_at,
            "body_status": self.body_status,
            "verification_status": self.verification_status,
            "metadata": dict(self.metadata or {}),
        }


def _acquisition_record(
    rule: ETFSourceRule,
    context: dict[str, Any],
    *,
    status: str,
    source_id: str | None = None,
    error: str | None = None,
) -> dict[str, Any]:
    """Return a stable record of the rule actually attempted for a snapshot."""

    result = rule.public(context)
    result.update({"status": status, "source_id": source_id})
    if error:
        result["error"] = str(error)[:240]
    return result


class OfficialETFProductProvider:
    """Fetch official manager/index/exchange inputs and normalize product fields."""

    def __init__(
        self,
        *,
        http_get: Callable[..., Any] = requests.get,
        timeout: int = 25,
        source_registry: ETFSourceRegistry | None = None,
    ) -> None:
        self.http_get = http_get
        self.timeout = timeout
        self.mapper = AuditedETFIndexMapper()
        self.source_registry = source_registry or get_etf_source_registry()

    def _get(self, url: str, *, params: dict[str, Any] | None = None) -> Any:
        headers = dict(_HEADERS)
        if "sse.com.cn" in url:
            headers["Referer"] = "https://www.sse.com.cn/"
        if "szse.cn" in url:
            headers["Referer"] = "https://www.szse.cn/"
            headers["Accept"] = "application/json,text/plain,*/*"
        for attempt in range(2):
            try:
                response = self.http_get(
                    url, params=params, headers=headers, timeout=self.timeout
                )
                response.raise_for_status()
                return response
            except requests.RequestException:
                if attempt == 1:
                    raise
        raise RuntimeError("official ETF source request failed without a response")

    def _html_source(
        self,
        *,
        rule: ETFSourceRule,
        context: dict[str, Any],
    ) -> _FetchedSource:
        url = rule.resolved_url(context)
        response = self._get(url)
        declared = getattr(response, "encoding", None)
        text, encoding = decode_official_html(response.content, declared)
        content = _plain_html(text)
        digest = _hash_bytes(response.content)
        return _FetchedSource(
            source_id=f"etfsource_{digest[:24]}", kind=rule.source_kind,
            title=rule.resolved_title(context), publisher=rule.publisher,
            url=url, content=content, content_hash=digest,
            retrieved_at=utc_now(), published_at=rule.published_at,
            verification_status=rule.verification_status,
            metadata={
                "original_encoding": encoding, "raw_sha256": digest,
                "source_rule_id": rule.rule_id, "parser_id": rule.parser_id,
                "registry_version": self.source_registry.schema_version,
            },
        )

    def _pdf_source(
        self,
        *,
        rule: ETFSourceRule,
        context: dict[str, Any],
    ) -> _FetchedSource:
        url = rule.resolved_url(context)
        response = self._get(url)
        content = extract_pdf_text(response.content)
        digest = _hash_bytes(response.content)
        return _FetchedSource(
            source_id=f"etfsource_{digest[:24]}", kind=rule.source_kind,
            title=rule.resolved_title(context), publisher=rule.publisher,
            url=url, content=content, content_hash=digest,
            retrieved_at=utc_now(), published_at=rule.published_at,
            verification_status=rule.verification_status,
            metadata={
                "raw_sha256": digest, "pdf_text_chars": len(content),
                "source_rule_id": rule.rule_id, "parser_id": rule.parser_id,
                "registry_version": self.source_registry.schema_version,
            },
        )

    def _sse_list_source(
        self,
        *,
        rule: ETFSourceRule,
        context: dict[str, Any],
    ) -> tuple[_FetchedSource, list[dict[str, Any]]]:
        response = self._get(rule.resolved_url(context), params=rule.resolved_params(context))
        payload = response.json()
        rows = [dict(item) for item in payload.get("result") or [] if isinstance(item, dict)]
        content = json.dumps(rows, ensure_ascii=False, sort_keys=True)
        digest = _hash_text(content)
        source = _FetchedSource(
            source_id=f"etfsource_{digest[:24]}", kind=rule.source_kind,
            title=rule.resolved_title(context), publisher=rule.publisher,
            url="https://www.sse.com.cn/assortment/fund/etf/list/", content=content,
            content_hash=digest, retrieved_at=utc_now(),
            verification_status=rule.verification_status,
            metadata={
                "row_count": len(rows), "sql_id": "FUND_LIST",
                "source_rule_id": rule.rule_id, "parser_id": rule.parser_id,
                "registry_version": self.source_registry.schema_version,
            },
        )
        return source, rows

    def _szse_list_source(
        self,
        *,
        rule: ETFSourceRule,
        context: dict[str, Any],
    ) -> tuple[_FetchedSource, list[dict[str, Any]]]:
        response = self._get(rule.resolved_url(context), params=rule.resolved_params(context))
        payload = response.json()
        rows: list[dict[str, Any]] = []
        catalog_data_as_of = ""
        for block in payload if isinstance(payload, list) else []:
            if not isinstance(block, dict):
                continue
            metadata = dict(block.get("metadata") or {})
            if str(metadata.get("tabkey") or "") != "tab1":
                continue
            catalog_data_as_of = str(metadata.get("subname") or "")
            rows.extend(
                {
                    **dict(item),
                    "_catalog_data_as_of": catalog_data_as_of,
                }
                for item in block.get("data") or []
                if isinstance(item, dict)
            )
        content = json.dumps(rows, ensure_ascii=False, sort_keys=True)
        digest = _hash_text(content)
        source = _FetchedSource(
            source_id=f"etfsource_{digest[:24]}", kind=rule.source_kind,
            title=rule.resolved_title(context), publisher=rule.publisher,
            url="https://www.szse.cn/www/market/product/list/etfList/",
            content=content, content_hash=digest, retrieved_at=utc_now(),
            published_at=catalog_data_as_of or None,
            verification_status=rule.verification_status,
            metadata={
                "row_count": len(rows), "catalog_id": "1945", "tab_key": "tab1",
                "source_rule_id": rule.rule_id, "parser_id": rule.parser_id,
                "registry_version": self.source_registry.schema_version,
            },
        )
        return source, rows

    def _fetch_rule(
        self,
        rule: ETFSourceRule,
        context: dict[str, Any],
    ) -> _FetchedSource:
        if rule.response_type == "html":
            return self._html_source(rule=rule, context=context)
        if rule.response_type == "pdf":
            return self._pdf_source(rule=rule, context=context)
        raise ValueError(
            f"unsupported product-profile response type {rule.response_type}: {rule.rule_id}"
        )

    @staticmethod
    def _fee_values(text: str) -> tuple[float | None, float | None]:
        management = re.search(r"管理费\s*([0-9.]+)%", text)
        custody = re.search(r"托管费\s*([0-9.]+)%", text)
        return (
            float(management.group(1)) / 100 if management else None,
            float(custody.group(1)) / 100 if custody else None,
        )

    @staticmethod
    def _pcf_values(text: str) -> dict[str, Any]:
        nav_match = re.search(
            r"(20\d{2}-\d{2}-\d{2})日\s*信息内容.*?基金份额净值\(单位：元\)\s*([0-9.]+)",
            text,
            re.S,
        )
        return {
            "data_as_of": nav_match.group(1) if nav_match else None,
            "unit_nav": _number(nav_match.group(2)) if nav_match else None,
            "iopv_published": "是否需要公布IOPV\n是" in text or "是否需要公布IOPV 是" in text,
        }

    @staticmethod
    def _methodology_values(text: str) -> dict[str, Any]:
        """Parse stable fields from a CSI methodology PDF's extracted text."""

        values: dict[str, Any] = {}
        version = re.search(r"版本号\s*([A-Z]?\d+(?:\.\d+)*)", text, re.I)
        published = re.search(r"(20\d{2})\s*年\s*(\d{1,2})\s*月", text)
        index_code = re.search(r"指数代码[：:]\s*(\d{6})", text)
        index_name = re.search(r"指数名称[：:]\s*([^\n\r]+)", text)
        component_count = re.search(r"(?:不超过|排名靠前的?)\s*(\d+)\s*(?:家|只).*?指数样本", text, re.S)
        weight_caps = [
            float(item) / 100
            for item in re.findall(r"单个样本权重不超过\s*([0-9.]+)%", text)
        ]
        rebalance_cap = re.search(r"每次调整的样本比例一般不超过\s*([0-9.]+)%", text)

        if version:
            raw_version = version.group(1).upper()
            values["version"] = raw_version if raw_version.startswith("V") else f"V{raw_version}"
        if published:
            values["published_at"] = f"{published.group(1)}-{int(published.group(2)):02d}"
        if index_code:
            values["index_code"] = f"{index_code.group(1)}.CSI"
        if index_name:
            values["index_name"] = index_name.group(1).strip()
        if component_count:
            values["target_component_count"] = int(component_count.group(1))
        if weight_caps:
            values["single_constituent_weight_cap"] = max(weight_caps)
            if len(set(weight_caps)) > 1:
                values["special_constituent_weight_cap"] = min(weight_caps)
        if "每半年调整一次" in text:
            values["review_frequency"] = "semiannual"
            values["review_months"] = [6, 12]
        elif "每季度调整一次" in text:
            values["review_frequency"] = "quarterly"
            values["review_months"] = [3, 6, 9, 12]
        elif "每年调整一次" in text:
            values["review_frequency"] = "annual"
        if rebalance_cap:
            values["regular_rebalance_change_cap"] = float(rebalance_cap.group(1)) / 100
        return values

    def fetch(
        self,
        symbol: str,
        *,
        as_of: str | None = None,
        instrument_profile: dict[str, Any] | None = None,
        universe_snapshot: ETFResearchSnapshot | None = None,
    ) -> dict[str, Any]:
        normalized = normalize_etf_symbol(symbol)
        retrieved_at = utc_now()
        errors: list[dict[str, str]] = []
        fetched: dict[str, _FetchedSource] = {}
        sse_rows: list[dict[str, Any]] = []
        szse_rows: list[dict[str, Any]] = []
        mapping_payload = dict((universe_snapshot.payload if universe_snapshot else {}).get("mapping") or {})
        tracked_code = str(
            mapping_payload.get("tracked_index_code")
            or mapping_payload.get("index_code")
            or ""
        ).upper()
        tracked_name = str(
            mapping_payload.get("tracked_index_name")
            or mapping_payload.get("index_name")
            or ""
        )
        audited = _AUDITED_588870 if normalized == "588870.SH" else {}
        audited_identity = dict(audited.get("identity") or {})
        acquisition: list[dict[str, Any]] = []
        bootstrap_context = source_context(
            normalized,
            manager=str(audited_identity.get("manager") or ""),
            index_code=tracked_code or str(audited_identity.get("tracked_index_code") or ""),
        )
        catalog_rules = self.source_registry.select(
            phase="product_profile",
            context=bootstrap_context,
            slots=("sse_catalog", "szse_catalog"),
        )
        for catalog_rule in catalog_rules:
            try:
                if catalog_rule.slot == "sse_catalog":
                    catalog_source, sse_rows = self._sse_list_source(
                        rule=catalog_rule, context=bootstrap_context
                    )
                else:
                    catalog_source, szse_rows = self._szse_list_source(
                        rule=catalog_rule, context=bootstrap_context
                    )
                fetched[catalog_rule.slot] = catalog_source
                acquisition.append(_acquisition_record(
                    catalog_rule, bootstrap_context, status="completed",
                    source_id=catalog_source.source_id,
                ))
            except Exception as exc:
                errors.append({"source": catalog_rule.rule_id, "error": str(exc)[:240]})
                acquisition.append(_acquisition_record(
                    catalog_rule, bootstrap_context, status="failed", error=str(exc),
                ))
        mapping_source_ids: list[str] = []
        sse_row = next(
            (row for row in sse_rows if str(row.get("fundCode") or "") == normalized[:6]),
            None,
        )
        szse_row = next(
            (
                row for row in szse_rows
                if re.search(
                    rf"(?<!\d){re.escape(normalized[:6])}(?!\d)",
                    _plain_html(str(row.get("sys_key") or "")),
                )
            ),
            None,
        )
        if sse_row:
            row_content = json.dumps(sse_row, ensure_ascii=False, sort_keys=True)
            row_digest = _hash_text(row_content)
            fetched["sse_subject"] = _FetchedSource(
                source_id=f"etfsource_{row_digest[:24]}", kind="fund_product",
                title=f"上海证券交易所 ETF 产品资料 {normalized[:6]}",
                publisher="上海证券交易所",
                url=(
                    "https://www.sse.com.cn/assortment/fund/list/etfinfo/basic/"
                    f"index.shtml?FUNDID={normalized[:6]}"
                ),
                content=row_content, content_hash=row_digest, retrieved_at=retrieved_at,
                metadata={
                    "fund_code": normalized[:6], "sql_id": "FUND_LIST",
                    "derived_from_rule_id": (
                        catalog_rules[0].rule_id if catalog_rules else None
                    ),
                    "registry_version": self.source_registry.schema_version,
                },
            )
            mapping_source_ids.append(fetched["sse_subject"].source_id)
            sse_code = str(sse_row.get("INDEX_CODE") or "").strip()
            if sse_code:
                tracked_code = (
                    tracked_code
                    if "." in tracked_code and _compact_index_code(tracked_code) == sse_code
                    else (
                        f"{sse_code}.CSI"
                        if sse_code.startswith("93")
                        else f"{sse_code}.SH"
                    )
                )
            tracked_name = str(sse_row.get("INDEX_NAME") or tracked_name)
        elif szse_row:
            row_content = json.dumps(szse_row, ensure_ascii=False, sort_keys=True)
            row_digest = _hash_text(row_content)
            fetched["szse_subject"] = _FetchedSource(
                source_id=f"etfsource_{row_digest[:24]}", kind="fund_product",
                title=f"深圳证券交易所 ETF 产品资料 {normalized[:6]}",
                publisher="深圳证券交易所",
                url="https://www.szse.cn/www/market/product/list/etfList/",
                content=row_content, content_hash=row_digest, retrieved_at=retrieved_at,
                published_at=str(szse_row.get("_catalog_data_as_of") or "") or None,
                metadata={
                    "fund_code": normalized[:6], "catalog_id": "1945", "tab_key": "tab1",
                    "derived_from_rule_id": (
                        next(
                            (rule.rule_id for rule in catalog_rules if rule.slot == "szse_catalog"),
                            None,
                        )
                    ),
                    "registry_version": self.source_registry.schema_version,
                },
            )
            mapping_source_ids.append(fetched["szse_subject"].source_id)
            index_cell = _plain_html(str(szse_row.get("nhzs") or ""))
            index_match = re.match(r"([0-9A-Z]{6,})\s*(.*)", index_cell, re.I)
            if index_match:
                szse_code = index_match.group(1).upper()
                tracked_code = (
                    tracked_code
                    if _compact_index_code(tracked_code) == szse_code
                    else _qualify_index_code(szse_code)
                )
                tracked_name = index_match.group(2).strip() or tracked_name
        if not tracked_code and self.mapper.supports(normalized):
            mapped = self.mapper.resolve(normalized, as_of=as_of).to_dict()
            tracked_code = str(mapped.get("tracked_index_code") or "")
            tracked_name = str(mapped.get("tracked_index_name") or "")

        manager = (
            str(sse_row.get("companyName") or "")
            if sse_row else (
                _plain_html(str(szse_row.get("glrmc") or ""))
                if szse_row else str(audited_identity.get("manager") or "")
            )
        )
        context = source_context(
            normalized,
            manager=manager,
            index_code=tracked_code or str(audited_identity.get("tracked_index_code") or ""),
        )
        enrichment_rules = [
            rule for rule in self.source_registry.select(
                phase="product_profile", context=context
            )
            if rule.slot not in {"sse_catalog", "szse_catalog"}
        ]
        with ThreadPoolExecutor(max_workers=min(4, len(enrichment_rules) or 1)) as executor:
            futures = {
                executor.submit(self._fetch_rule, rule, context): rule
                for rule in enrichment_rules
            }
            for future in as_completed(futures):
                rule = futures[future]
                try:
                    source = future.result()
                    fetched[rule.slot] = source
                    acquisition.append(_acquisition_record(
                        rule, context, status="completed", source_id=source.source_id,
                    ))
                except Exception as exc:
                    errors.append({"source": rule.rule_id, "error": str(exc)[:240]})
                    acquisition.append(_acquisition_record(
                        rule, context, status="failed", error=str(exc),
                    ))

        identity_source_ids = list(mapping_source_ids)
        if fetched.get("annual_report"):
            identity_source_ids.append(fetched["annual_report"].source_id)
        identity_as_of = (
            "2025-12-31"
            if normalized == "588870.SH"
            else str(
                (szse_row or {}).get("_catalog_data_as_of")
                or as_of
                or retrieved_at
            )
        )
        identity_values = {
            **audited_identity,
            "fund_short_name": (
                str(sse_row.get("secNameFull") or sse_row.get("fundAbbr") or "")
                if sse_row else (
                    _plain_html(str(szse_row.get("kzjcurl") or ""))
                    if szse_row else audited_identity.get("fund_short_name")
                )
            ),
            "manager": (
                str(sse_row.get("companyName") or "")
                if sse_row else (
                    _plain_html(str(szse_row.get("glrmc") or ""))
                    if szse_row else audited_identity.get("manager")
                )
            ),
            "custodian": str(sse_row.get("TRUSTEE_NAME") or "") if sse_row else audited_identity.get("custodian"),
            "listing_date": _iso_date(sse_row.get("listingDate")) if sse_row else audited_identity.get("listing_date"),
            "exchange": "上海证券交易所" if normalized.endswith(".SH") else "深圳证券交易所",
            "tracked_index_code": tracked_code or audited_identity.get("tracked_index_code"),
            "tracked_index_name": tracked_name or audited_identity.get("tracked_index_name"),
        }
        identity = {
            key: _field(
                value,
                source_ids=identity_source_ids,
                data_as_of=identity_as_of,
                semantics=f"official_fund_identity.{key}",
            )
            for key, value in identity_values.items()
            if key != "exchange_short_name"
        }
        conflicts: list[dict[str, Any]] = []
        if sse_row and audited_identity:
            comparisons = {
                "manager": str(sse_row.get("companyName") or ""),
                "custodian": str(sse_row.get("TRUSTEE_NAME") or ""),
                "listing_date": str(_iso_date(sse_row.get("listingDate")) or ""),
                "tracked_index_code": tracked_code,
                "tracked_index_name": tracked_name,
            }
            for key, official_value in comparisons.items():
                audited_value = str(audited_identity.get(key) or "")
                normalized_official = re.sub(r"\s+", "", official_value)
                normalized_audited = re.sub(r"\s+", "", audited_value)
                if normalized_official and normalized_audited and normalized_official != normalized_audited:
                    conflicts.append({
                        "field": key, "official_exchange_value": official_value,
                        "audited_manager_value": audited_value,
                        "source_ids": identity_source_ids,
                    })
                    identity[key] = {**identity[key], "status": "conflict", "note": "官方来源值不一致"}

        methodology_values = dict(audited.get("index_methodology") or {})
        if fetched.get("methodology"):
            methodology_values.update(
                self._methodology_values(fetched["methodology"].content)
            )
            methodology_values["source_url"] = fetched["methodology"].url
        methodology_source_ids = [fetched["methodology"].source_id] if fetched.get("methodology") else []
        methodology_as_of = str(methodology_values.get("published_at") or as_of or retrieved_at)
        methodology = {
            key: _field(
                value,
                source_ids=methodology_source_ids,
                data_as_of=methodology_as_of,
                semantics=f"official_index_methodology.{key}",
                unit=_FIELD_UNITS.get(key),
            )
            for key, value in methodology_values.items()
            if key != "source_url"
        }
        methodology["source_url"] = _field(
            methodology_values.get("source_url"), source_ids=methodology_source_ids,
            data_as_of=methodology_as_of, semantics="official_index_methodology.original_url",
        )
        for identity_key, methodology_key in (
            ("tracked_index_code", "index_code"),
            ("tracked_index_name", "index_name"),
        ):
            current = identity.get(identity_key) or {}
            replacement = methodology.get(methodology_key) or {}
            placeholder_name = (
                identity_key == "tracked_index_name"
                and _compact_index_name(current.get("value"))
                == _compact_index_code((identity.get("tracked_index_code") or {}).get("value"))
            )
            if (
                current.get("status") != "available" or placeholder_name
            ) and replacement.get("status") == "available":
                identity[identity_key] = {
                    **replacement,
                    "semantics": f"official_index_methodology.{identity_key}",
                    "note": "交易所产品目录未提供完整值，使用官方指数编制方案补齐",
                }

        metric_values: dict[str, tuple[Any, str, list[str], str]] = {}
        if normalized == "588870.SH":
            annual_id = [fetched["annual_report"].source_id] if fetched.get("annual_report") else []
            for key, value in dict(audited.get("reported_metrics") or {}).items():
                metric_values[key] = (value, str(audited.get("reported_as_of")), annual_id, f"annual_report.{key}")
            product = fetched.get("product")
            if product:
                management_fee, custody_fee = self._fee_values(product.content)
                metric_values["management_fee_rate"] = (
                    management_fee, product.retrieved_at[:10], [product.source_id], "contractual_annual_rate"
                )
                metric_values["custody_fee_rate"] = (
                    custody_fee, product.retrieved_at[:10], [product.source_id], "contractual_annual_rate"
                )
            pcf = fetched.get("pcf")
            if pcf:
                pcf_values = self._pcf_values(pcf.content)
                metric_values["unit_nav"] = (
                    pcf_values.get("unit_nav"), str(pcf_values.get("data_as_of") or pcf.retrieved_at[:10]),
                    [pcf.source_id], "fund_share_nav_from_pcf",
                )
                metric_values["iopv"] = (
                    None, str(pcf_values.get("data_as_of") or pcf.retrieved_at[:10]),
                    [pcf.source_id], "iopv_value_not_published_on_source_page",
                )

        metrics_by_key = _metric_map(instrument_profile)
        instrument_sources = {
            str(item.get("source_id") or ""): dict(item)
            for item in ((instrument_profile or {}).get("sources") or [])
            if isinstance(item, dict) and item.get("source_id")
        }
        for target_key, profile_key, semantics in (
            ("fund_units", "total_shares", "listed_fund_units"),
            ("exchange_market_value", "total_market_cap", "market_price_times_listed_fund_units"),
        ):
            raw = metrics_by_key.get(profile_key) or {}
            source_id = str(raw.get("source_id") or "")
            metric_values[target_key] = (
                raw.get("value") if raw.get("status") == "available" else None,
                str(raw.get("data_as_of") or (instrument_profile or {}).get("data_as_of") or retrieved_at),
                [source_id] if source_id else [], semantics,
            )
            if source_id and source_id in instrument_sources:
                source = instrument_sources[source_id]
                digest = _hash_text(json.dumps(source, ensure_ascii=False, sort_keys=True))
                fetched[f"instrument:{source_id}"] = _FetchedSource(
                    source_id=source_id, kind="market_data",
                    title=str(source.get("label") or "ETF 行情与份额快照"),
                    publisher=str(source.get("provider_id") or "行情提供方"),
                    url=str(source.get("url") or f"provider://{source_id}"),
                    content=json.dumps(source, ensure_ascii=False, sort_keys=True),
                    content_hash=digest, retrieved_at=str(source.get("retrieved_at") or retrieved_at),
                    published_at=str(source.get("data_as_of") or "") or None,
                    verification_status="source_recorded", body_status="structured_payload",
                )

        nav = _number((metric_values.get("unit_nav") or (None,))[0])
        current_price = _number((metrics_by_key.get("current_price") or {}).get("value"))
        price_date = str((metrics_by_key.get("current_price") or {}).get("data_as_of") or "")[:10]
        nav_date = str((metric_values.get("unit_nav") or (None, "", [], ""))[1])[:10]
        premium = None
        premium_note = "价格和净值/IOPV不属于同一交易日，未计算折溢价"
        premium_sources: list[str] = []
        if current_price is not None and nav not in {None, 0} and price_date and price_date == nav_date:
            premium = current_price / nav - 1
            premium_note = "同一交易日市场价格相对基金份额净值"
            premium_sources = [
                str((metrics_by_key.get("current_price") or {}).get("source_id") or ""),
                *((metric_values.get("unit_nav") or (None, "", [], ""))[2]),
            ]
        metric_values["premium_discount_rate"] = (
            premium, nav_date or price_date or retrieved_at[:10], premium_sources,
            "same_day_market_price_divided_by_nav_minus_one",
        )
        product_metrics = {
            key: _field(
                value, source_ids=source_ids, data_as_of=data_as_of,
                semantics=semantics, unit=_FIELD_UNITS.get(key),
                note=(premium_note if key == "premium_discount_rate" else None),
                formula=(
                    "current_price / unit_nav - 1"
                    if key == "premium_discount_rate" and value is not None
                    else "fund_units * current_price"
                    if key == "exchange_market_value"
                    and value is not None
                    and str(semantics).startswith("market_price_times_")
                    else None
                ),
                input_metrics=(
                    ("current_price", "unit_nav")
                    if key == "premium_discount_rate" and value is not None
                    else ("fund_units", "current_price")
                    if key == "exchange_market_value"
                    and value is not None
                    and str(semantics).startswith("market_price_times_")
                    else ()
                ),
                calculation_version=(
                    "etf-product-calc-v1"
                    if key in {"premium_discount_rate", "exchange_market_value"}
                    and value is not None
                    else None
                ),
            )
            for key, (value, data_as_of, source_ids, semantics) in metric_values.items()
        }

        hard_keys = (
            (identity, "manager"), (identity, "exchange"),
            (identity, "tracked_index_code"), (identity, "tracked_index_name"),
            (methodology, "version"), (methodology, "source_url"),
        )
        missing_hard = [
            key for section, key in hard_keys
            if (section.get(key) or {}).get("status") != "available"
            or not (section.get(key) or {}).get("source_ids")
        ]
        optional_keys = (
            "management_fee_rate", "custody_fee_rate", "unit_nav", "fund_units",
            "published_net_assets", "exchange_market_value", "iopv", "premium_discount_rate",
        )
        missing_optional = [
            key for key in optional_keys
            if (product_metrics.get(key) or {}).get("status") != "available"
        ]
        return {
            "symbol": normalized,
            "data_as_of": max(
                [
                    str(item.get("data_as_of") or "")
                    for item in product_metrics.values()
                    if isinstance(item, dict)
                ] or [str(as_of or retrieved_at)]
            ),
            "retrieved_at": retrieved_at,
            "identity": identity,
            "index_methodology": methodology,
            "product_metrics": product_metrics,
            "sources": [item.public() for item in fetched.values()],
            "hard_gate_status": "passed" if not missing_hard else "failed_validation",
            "quality_status": (
                "failed_validation" if missing_hard else
                "passed_with_gaps" if missing_optional or errors else "passed"
            ),
            "missing_hard_fields": missing_hard,
            "missing_optional_fields": missing_optional,
            "conflicts": conflicts,
            "source_errors": errors,
            "source_acquisition": {
                "registry_version": self.source_registry.schema_version,
                "rules": sorted(acquisition, key=lambda item: str(item.get("rule_id") or "")),
            },
        }


class ETFShareFlowProvider:
    """Track official exchange fund units and same-index peer changes."""

    def __init__(
        self,
        *,
        http_get: Callable[..., Any] = requests.get,
        timeout: int = 20,
        source_registry: ETFSourceRegistry | None = None,
    ) -> None:
        self.http_get = http_get
        self.timeout = timeout
        self.source_registry = source_registry or get_etf_source_registry()

    def _get(self, url: str, *, params: dict[str, Any]) -> Any:
        headers = dict(_HEADERS)
        headers["Referer"] = (
            "https://www.sse.com.cn/" if "sse.com.cn" in url
            else "https://fund.szse.cn/marketdata/fundslist/index.html"
        )
        response = self.http_get(url, params=params, headers=headers, timeout=self.timeout)
        response.raise_for_status()
        return response

    def _sse_catalog(
        self, context: dict[str, Any]
    ) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        rule = self.source_registry.get("sse.etf_peer_catalog.v1")
        response = self._get(
            rule.resolved_url(context), params=rule.resolved_params(context)
        )
        rows = [dict(item) for item in response.json().get("result") or [] if isinstance(item, dict)]
        content = json.dumps(rows, ensure_ascii=False, sort_keys=True)
        digest = _hash_text(content)
        return rows, {
            "source_id": f"etfsource_{digest[:24]}", "kind": rule.source_kind,
            "title": rule.resolved_title(context), "publisher": rule.publisher,
            "url": "https://www.sse.com.cn/assortment/fund/etf/list/",
            "content_hash": digest, "retrieved_at": utc_now(),
            "verification_status": rule.verification_status, "body_status": "full_text",
            "content": content, "metadata": {
                "source_rule_id": rule.rule_id, "parser_id": rule.parser_id,
                "registry_version": self.source_registry.schema_version,
            },
        }

    def _sse_scale(self, day: date) -> tuple[str, list[dict[str, Any]], dict[str, Any]]:
        day_text = day.isoformat()
        rule = self.source_registry.get("sse.etf_share_scale_history.v1")
        context = {"data_as_of": day_text}
        response = self._get(
            rule.resolved_url(context), params=rule.resolved_params(context)
        )
        rows = [dict(item) for item in response.json().get("result") or [] if isinstance(item, dict)]
        content = json.dumps(rows, ensure_ascii=False, sort_keys=True)
        digest = _hash_text(content)
        source = {
            "source_id": f"etfsource_{digest[:24]}", "kind": rule.source_kind,
            "title": rule.resolved_title(context),
            "publisher": rule.publisher,
            "url": "https://www.sse.com.cn/market/funddata/volumn/etfvolumn/",
            "content_hash": digest, "retrieved_at": utc_now(), "published_at": day_text,
            "verification_status": rule.verification_status, "body_status": "full_text",
            "content": content, "metadata": {
                "source_rule_id": rule.rule_id, "parser_id": rule.parser_id,
                "registry_version": self.source_registry.schema_version,
            },
        }
        return day_text, rows, source

    def _szse_current(
        self, context: dict[str, Any]
    ) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        rule = self.source_registry.get("szse.etf_share_scale_current.v1")
        response = self._get(
            rule.resolved_url(context), params=rule.resolved_params(context),
        )
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(iterator)]
        rows = [
            {headers[index]: value for index, value in enumerate(values)}
            for values in iterator
            if values and any(value is not None for value in values)
        ]
        digest = _hash_bytes(response.content)
        source = {
            "source_id": f"etfsource_{digest[:24]}", "kind": rule.source_kind,
            "title": rule.resolved_title(context), "publisher": rule.publisher,
            "url": "https://fund.szse.cn/marketdata/fundslist/index.html",
            "content_hash": digest, "retrieved_at": utc_now(),
            "verification_status": rule.verification_status, "body_status": "full_text",
            "content": json.dumps(rows, ensure_ascii=False, sort_keys=True, default=str),
            "metadata": {
                "source_rule_id": rule.rule_id, "parser_id": rule.parser_id,
                "registry_version": self.source_registry.schema_version,
            },
        }
        return rows, source

    def _prices(
        self, symbols: list[str], context: dict[str, Any]
    ) -> tuple[dict[str, float], dict[str, Any] | None]:
        sh_symbols = {symbol for symbol in symbols if symbol.endswith(".SH")}
        if not sh_symbols:
            return {}, None
        rule = self.source_registry.get("sse.etf_market_price.v1")
        response = self._get(
            rule.resolved_url(context), params=rule.resolved_params(context),
        )
        payload = response.json()
        rows = [item for item in payload.get("list") or [] if isinstance(item, list)]
        prices = {
            f"{str(row[0])}.SH": float(row[2])
            for row in rows
            if len(row) >= 3 and f"{str(row[0])}.SH" in sh_symbols and _number(row[2]) is not None
        }
        content = json.dumps(rows, ensure_ascii=False, sort_keys=True)
        digest = _hash_text(content)
        return prices, {
            "source_id": f"etfsource_{digest[:24]}", "kind": rule.source_kind,
            "title": rule.resolved_title(context), "publisher": rule.publisher,
            "url": "https://www.sse.com.cn/assortment/fund/etf/price/",
            "content_hash": digest, "retrieved_at": utc_now(),
            "published_at": str(payload.get("date") or ""),
            "verification_status": rule.verification_status,
            "body_status": "full_text", "content": content,
            "metadata": {
                "source_rule_id": rule.rule_id, "parser_id": rule.parser_id,
                "registry_version": self.source_registry.schema_version,
            },
        }

    def fetch(
        self,
        symbol: str,
        *,
        tracked_index_code: str,
        tracked_index_name: str,
        as_of: str | None = None,
        previous_peer_group: ETFResearchSnapshot | None = None,
    ) -> dict[str, Any]:
        normalized = normalize_etf_symbol(symbol)
        target_day = _as_of_date(as_of)
        context = {
            **source_context(normalized, index_code=tracked_index_code),
            "data_as_of": target_day.isoformat(),
        }
        acquisition: list[dict[str, Any]] = []
        catalog_rule = self.source_registry.get("sse.etf_peer_catalog.v1")
        catalog, catalog_source = self._sse_catalog(context)
        acquisition.append(_acquisition_record(
            catalog_rule, context, status="completed",
            source_id=str(catalog_source.get("source_id") or ""),
        ))
        index_number = str(tracked_index_code or "").split(".", 1)[0]
        sse_peers = [
            {
                "symbol": f"{str(row.get('fundCode'))}.SH",
                "name": str(row.get("secNameFull") or row.get("fundAbbr") or ""),
                "manager": str(row.get("companyName") or ""),
                "custodian": str(row.get("TRUSTEE_NAME") or ""),
                "listing_date": _iso_date(row.get("listingDate")),
                "mapping_status": "official_index_code",
                "mapping_source_ids": [catalog_source["source_id"]],
            }
            for row in catalog
            if str(row.get("INDEX_CODE") or "").strip() == index_number
        ]

        sources = [catalog_source]
        if previous_peer_group is not None:
            sources.extend(
                dict(item)
                for item in previous_peer_group.payload.get("_source_catalog") or []
                if isinstance(item, dict) and item.get("source_id")
            )
        errors: list[str] = []
        szse_peers: list[dict[str, Any]] = []
        szse_rule = self.source_registry.get("szse.etf_share_scale_current.v1")
        try:
            szse_rows, szse_source = self._szse_current(context)
            sources.append(szse_source)
            acquisition.append(_acquisition_record(
                szse_rule, context, status="completed",
                source_id=str(szse_source.get("source_id") or ""),
            ))
            aliases = _index_name_aliases(tracked_index_code, tracked_index_name)
            for row in szse_rows:
                name = str(_row_value(row, _SZSE_FUND_NAME_FIELDS) or "")
                disclosed_index_code = _row_value(row, _SZSE_INDEX_CODE_FIELDS)
                disclosed_index_name = _row_value(row, _SZSE_INDEX_NAME_FIELDS)
                if disclosed_index_code is not None:
                    if not _same_index_code(disclosed_index_code, tracked_index_code):
                        continue
                    mapping_status = "official_index_code"
                elif disclosed_index_name is not None:
                    if not _same_index_name(disclosed_index_name, aliases):
                        continue
                    mapping_status = "official_index_name"
                elif aliases and _fund_name_matches_index(name, aliases):
                    mapping_status = "name_alias_requires_cross_check"
                else:
                    continue
                raw_fund_code = str(_row_value(row, _SZSE_FUND_CODE_FIELDS) or "").strip()
                if not raw_fund_code.isdigit() or len(raw_fund_code) > 6:
                    continue
                fund_code = raw_fund_code.zfill(6)
                szse_peers.append({
                    "symbol": f"{fund_code}.SZ",
                    "name": name,
                    "manager": str(row.get("基金管理人") or ""),
                    "custodian": str(row.get("基金托管人") or ""),
                    "listing_date": _iso_date(row.get("上市日期")),
                    "mapping_status": mapping_status,
                    "mapping_source_ids": [szse_source["source_id"]],
                    "current_units_from_szse": _number(
                        _row_value(row, _SZSE_FUND_UNITS_FIELDS)
                    ),
                    "current_market_price_from_szse": _number(
                        _row_value(row, _SZSE_MARKET_PRICE_FIELDS)
                    ),
                    "current_nav_from_szse": _number(
                        _row_value(row, _SZSE_NAV_FIELDS)
                    ),
                    "disclosed_index_code": (
                        str(disclosed_index_code) if disclosed_index_code is not None else None
                    ),
                    "disclosed_index_name": (
                        str(disclosed_index_name) if disclosed_index_name is not None else None
                    ),
                })
        except Exception as exc:
            errors.append(f"szse_current:{str(exc)[:180]}")
            acquisition.append(_acquisition_record(
                szse_rule, context, status="failed", error=str(exc),
            ))

        peers_by_symbol = {
            item["symbol"]: item for item in [*sse_peers, *szse_peers] if item["symbol"][:6].isdigit()
        }
        if normalized not in peers_by_symbol:
            peers_by_symbol[normalized] = {
                "symbol": normalized, "name": normalized, "manager": "", "custodian": "",
                "listing_date": None, "mapping_status": "subject_fallback",
                "mapping_source_ids": [],
            }

        candidate_days: list[date] = []
        cursor = target_day
        while len(candidate_days) < 12:
            if cursor.weekday() < 5:
                candidate_days.append(cursor)
            cursor -= timedelta(days=1)
        candidate_days.extend([target_day - timedelta(days=20), target_day - timedelta(days=35)])
        scale_results: list[tuple[str, list[dict[str, Any]], dict[str, Any]]] = []
        scale_errors = 0
        with ThreadPoolExecutor(max_workers=6) as executor:
            futures = {executor.submit(self._sse_scale, day): day for day in candidate_days}
            for future in as_completed(futures):
                try:
                    result = future.result()
                    if result[1]:
                        scale_results.append(result)
                except Exception as exc:
                    scale_errors += 1
                    errors.append(f"sse_scale:{futures[future].isoformat()}:{str(exc)[:120]}")
        scale_results.sort(key=lambda item: item[0], reverse=True)
        scale_rule = self.source_registry.get("sse.etf_share_scale_history.v1")
        scale_record = _acquisition_record(
            scale_rule,
            context,
            status=(
                "completed_with_gaps" if scale_results and scale_errors
                else "completed" if scale_results else "failed"
            ),
            source_id=(
                str(scale_results[0][2].get("source_id") or "")
                if scale_results else None
            ),
            error=(f"{scale_errors} 个候选交易日获取失败" if scale_errors else None),
        )
        scale_record["source_ids"] = sorted({
            str(item[2].get("source_id") or "") for item in scale_results
            if item[2].get("source_id")
        })
        scale_record["successful_observation_days"] = len(scale_results)
        acquisition.append(scale_record)
        series: dict[str, list[dict[str, Any]]] = {key: [] for key in peers_by_symbol}
        for day_text, rows, source in scale_results:
            sources.append(source)
            values = {
                f"{str(row.get('SEC_CODE') or '')}.SH": _number(row.get("TOT_VOL"))
                for row in rows
            }
            for peer_symbol in list(series):
                raw_units = values.get(peer_symbol)
                if raw_units is None:
                    continue
                # SSE reports TOT_VOL in 10,000 fund units.
                series[peer_symbol].append({
                    "data_as_of": day_text, "fund_units": raw_units * 10_000,
                    "source_ids": [source["source_id"]],
                })

        previous_members = {
            str(item.get("symbol") or ""): dict(item)
            for item in ((previous_peer_group.payload if previous_peer_group else {}).get("members") or [])
            if isinstance(item, dict)
        }
        latest_day = scale_results[0][0] if scale_results else target_day.isoformat()
        for peer in szse_peers:
            peer_symbol = peer["symbol"]
            current_day = target_day.isoformat()
            if peer.get("current_units_from_szse") is not None:
                series[peer_symbol].append({
                    "data_as_of": current_day,
                    "fund_units": peer["current_units_from_szse"],
                    "source_ids": peer["mapping_source_ids"],
                })
            previous = previous_members.get(peer_symbol)
            previous_observations = list((previous or {}).get("history") or [])
            if previous and previous.get("current_units") is not None and not previous_observations:
                previous_observations = [{
                    "data_as_of": previous.get("data_as_of"),
                    "fund_units": previous.get("current_units"),
                    "source_ids": list(previous.get("source_ids") or []),
                }]
            existing_days = {
                str(item.get("data_as_of") or "") for item in series[peer_symbol]
            }
            for observation in previous_observations:
                observation_day = str(observation.get("data_as_of") or "")
                if not observation_day or observation_day in existing_days:
                    continue
                series[peer_symbol].append({
                    "data_as_of": observation_day,
                    "fund_units": observation.get("fund_units"),
                    "source_ids": list(observation.get("source_ids") or []),
                    "from_previous_snapshot": True,
                })
                existing_days.add(observation_day)
            series[peer_symbol].sort(key=lambda item: str(item.get("data_as_of") or ""), reverse=True)

        try:
            prices, price_source = self._prices(sorted(peers_by_symbol), context)
            if price_source:
                sources.append(price_source)
                acquisition.append(_acquisition_record(
                    self.source_registry.get("sse.etf_market_price.v1"),
                    context, status="completed",
                    source_id=str(price_source.get("source_id") or ""),
                ))
        except Exception as exc:
            prices, price_source = {}, None
            errors.append(f"peer_prices:{str(exc)[:180]}")
            acquisition.append(_acquisition_record(
                self.source_registry.get("sse.etf_market_price.v1"),
                context, status="failed", error=str(exc),
            ))

        members: list[dict[str, Any]] = []
        for peer_symbol, peer in sorted(peers_by_symbol.items()):
            observations = series.get(peer_symbol) or []
            current = observations[0] if observations else {}
            current_units = _number(current.get("fund_units"))

            def delta_at(index: int) -> float | None:
                if current_units is None or len(observations) <= index:
                    return None
                previous_units = _number(observations[index].get("fund_units"))
                return current_units - previous_units if previous_units is not None else None

            delta_1d = delta_at(1)
            delta_5d = delta_at(min(5, len(observations) - 1)) if len(observations) > 1 else None
            delta_20d = delta_at(len(observations) - 1) if len(observations) > 1 else None
            market_price = prices.get(peer_symbol)
            if market_price is None:
                market_price = _number(peer.get("current_market_price_from_szse"))
            nav_proxy = _number(peer.get("current_nav_from_szse"))
            estimation_price = market_price if market_price is not None else nav_proxy
            estimation_price_type = (
                "exchange_market_price"
                if market_price is not None
                else "exchange_published_nav_proxy"
                if nav_proxy is not None
                else None
            )
            estimated_flow = (
                delta_1d * estimation_price
                if delta_1d is not None and estimation_price is not None
                else None
            )
            member_sources = sorted({
                *list(peer.get("mapping_source_ids") or []),
                *[sid for observation in observations[:2] for sid in observation.get("source_ids") or []],
                *([str(price_source.get("source_id"))] if price_source and prices.get(peer_symbol) is not None else []),
            })
            members.append({
                "symbol": peer_symbol, "name": peer.get("name"),
                "manager": peer.get("manager"), "custodian": peer.get("custodian"),
                "listing_date": peer.get("listing_date"),
                "mapping_status": peer.get("mapping_status"),
                "data_as_of": str(current.get("data_as_of") or latest_day),
                "current_units": current_units,
                "unit": "fund_units",
                "delta_1d": delta_1d, "delta_5d": delta_5d, "delta_20d": delta_20d,
                "current_price": market_price,
                "estimation_price": estimation_price,
                "estimation_price_type": estimation_price_type,
                "estimated_net_flow_1d": estimated_flow,
                "estimated_net_flow_semantics": (
                    "share_delta_times_current_market_price_proxy"
                    if estimation_price_type == "exchange_market_price"
                    else "share_delta_times_exchange_published_nav_proxy"
                    if estimation_price_type == "exchange_published_nav_proxy"
                    else "unavailable"
                ),
                "source_ids": member_sources,
                "history": observations,
            })

        flows = [item["estimated_net_flow_1d"] for item in members if item["estimated_net_flow_1d"] is not None]
        deltas = [item["delta_1d"] for item in members if item["delta_1d"] is not None]
        return {
            "symbol": normalized, "tracked_index_code": tracked_index_code,
            "tracked_index_name": tracked_index_name, "data_as_of": latest_day,
            "retrieved_at": utc_now(), "members": members,
            "member_count": len(members),
            "official_index_mapping_count": sum(item["mapping_status"] == "official_index_code" for item in members),
            "name_mapped_count": sum(item["mapping_status"] == "name_alias_requires_cross_check" for item in members),
            "estimated_net_flow_1d": sum(flows) if flows else None,
            "inflow_member_ratio_1d": (
                sum(delta > 0 for delta in deltas) / len(deltas) if deltas else None
            ),
            "flow_coverage_ratio": len(flows) / len(members) if members else 0.0,
            "unit_change_coverage_ratio": len(deltas) / len(members) if members else 0.0,
            "estimated_net_flow_semantics": "sum_of_member_share_delta_times_available_exchange_price_or_nav_proxy",
            "market_price_flow_count": sum(
                item.get("estimated_net_flow_1d") is not None
                and item.get("estimation_price_type") == "exchange_market_price"
                for item in members
            ),
            "nav_proxy_flow_count": sum(
                item.get("estimated_net_flow_1d") is not None
                and item.get("estimation_price_type") == "exchange_published_nav_proxy"
                for item in members
            ),
            "sources": list({
                str(item.get("source_id") or ""): item
                for item in sources if item.get("source_id")
            }.values()),
            "warnings": [
                "深交所公开份额表不含跟踪指数代码；深市同组成员通过指数名称别名匹配，需交叉核验。"
            ] if any(
                item.get("mapping_status") == "name_alias_requires_cross_check"
                for item in szse_peers
            ) else [],
            "errors": errors,
            "source_acquisition": {
                "registry_version": self.source_registry.schema_version,
                "rules": sorted(acquisition, key=lambda item: str(item.get("rule_id") or "")),
            },
        }


class ETFProductProfileService:
    """Persist and project the canonical immutable ETF product profile."""

    def __init__(
        self,
        *,
        store: ETFResearchStore | None = None,
        provider: OfficialETFProductProvider | None = None,
        share_provider: ETFShareFlowProvider | None = None,
        ingestion: SourceIngestionService | None = None,
        source_registry: ETFSourceRegistry | None = None,
    ) -> None:
        self.store = store or ETFResearchStore()
        self.source_registry = (
            source_registry
            or getattr(provider, "source_registry", None)
            or getattr(share_provider, "source_registry", None)
            or get_etf_source_registry()
        )
        self.provider = provider or OfficialETFProductProvider(
            source_registry=self.source_registry
        )
        self.share_provider = share_provider or ETFShareFlowProvider(
            source_registry=self.source_registry
        )
        self.ingestion = ingestion

    def _ingestion(self) -> SourceIngestionService | None:
        if self.ingestion is not None:
            return self.ingestion
        try:
            self.ingestion = SourceIngestionService()
        except Exception:
            return None
        return self.ingestion

    def _archive_sources(self, symbol: str, sources: list[dict[str, Any]], origin_id: str) -> list[dict[str, Any]]:
        archived: list[dict[str, Any]] = []
        ingestion = self._ingestion()
        for raw in sources:
            item = dict(raw)
            content = str(item.pop("content", "") or "")
            if ingestion is not None and content:
                try:
                    observation = ingestion.ingest(
                        CollectedSource(
                            subject_key=symbol, source_kind=str(item.get("kind") or "fund_product"),
                            provider_id=str(item.get("publisher") or "official_etf_profile"),
                            publisher=str(item.get("publisher") or "官方来源"),
                            title=str(item.get("title") or "ETF 产品资料"),
                            source_locator=str(item.get("url") or ""), content=content,
                            published_at=item.get("published_at"),
                            retrieved_at=str(item.get("retrieved_at") or utc_now()),
                            verification_status=str(item.get("verification_status") or "source_recorded"),
                            body_status=str(item.get("body_status") or "full_text"),
                            source_class="company_disclosure",
                            metadata={
                                **dict(item.get("metadata") or {}),
                                "summary": item.get("title"),
                                "content_hash": item.get("content_hash"),
                            },
                        ),
                        origin_type="etf_product_profile", origin_id=origin_id,
                    )
                    item["document_ref"] = observation.get("document_ref")
                    item["verification_status"] = observation.get("verification_status") or item.get("verification_status")
                except Exception as exc:
                    item["archive_error"] = str(exc)[:180]
            archived.append(item)
        return archived

    @staticmethod
    def _section_coverage(section: dict[str, Any]) -> float:
        values = [item for item in section.values() if isinstance(item, dict) and "status" in item]
        return sum(item.get("status") == "available" for item in values) / len(values) if values else 0.0

    def _latest_section_with_fields(
        self,
        symbol: str,
        snapshot_type: str,
        required_fields: tuple[str, ...],
    ) -> ETFResearchSnapshot | None:
        for snapshot in self.store.recent_snapshots(symbol, snapshot_type, limit=50):
            if all(
                isinstance(snapshot.payload.get(key), dict)
                and snapshot.payload[key].get("status") == "available"
                and snapshot.payload[key].get("source_ids")
                for key in required_fields
            ):
                return snapshot
        return None

    @staticmethod
    def _mark_rule_cache_reuse(
        acquisition: dict[str, Any],
        *,
        slots: set[str],
        snapshot_id: str,
    ) -> None:
        for rule in acquisition.get("rules") or []:
            if rule.get("slot") not in slots or rule.get("status") != "failed":
                continue
            rule["status"] = "completed_with_gaps"
            rule["cache_reused"] = True
            rule["cache_snapshot_id"] = snapshot_id

    def _save_section(
        self,
        *,
        symbol: str,
        snapshot_type: str,
        data_as_of: str,
        payload: dict[str, Any],
        sources: list[dict[str, Any]],
        freshness_days: int,
    ) -> ETFResearchSnapshot:
        source_ids = sorted({str(item.get("source_id") or "") for item in sources if item.get("source_id")})
        stored_payload = {**payload, "_source_catalog": _stable_source_catalog(sources)}
        snapshot = build_etf_snapshot(
            symbol=symbol, snapshot_type=snapshot_type, data_as_of=data_as_of,
            payload=stored_payload, coverage_ratio=self._section_coverage(payload),
            source_ids=source_ids, retrieved_at=utc_now(),
            freshness_expires_at=(datetime.now(timezone.utc) + timedelta(days=freshness_days)).isoformat(),
            minimum_coverage=0.0,
        )
        return self.store.save_snapshot(snapshot)[0]

    def refresh(
        self,
        symbol: str,
        *,
        as_of: str | None = None,
        instrument_profile: dict[str, Any] | None = None,
        universe_snapshot: ETFResearchSnapshot | None = None,
        include_share_flows: bool = True,
    ) -> dict[str, Any]:
        normalized = normalize_etf_symbol(symbol)
        raw = self.provider.fetch(
            normalized, as_of=as_of, instrument_profile=instrument_profile,
            universe_snapshot=universe_snapshot,
        )
        cache_reused_sections: list[dict[str, str]] = []
        cached_source_catalog: list[dict[str, Any]] = []

        identity_required = (
            "manager", "exchange", "tracked_index_code", "tracked_index_name",
        )
        if any(
            (raw.get("identity", {}).get(key) or {}).get("status") != "available"
            or not (raw.get("identity", {}).get(key) or {}).get("source_ids")
            for key in identity_required
        ):
            cached_identity = self._latest_section_with_fields(
                normalized, "identity", identity_required
            )
            if cached_identity is not None:
                cached_source_catalog.extend(
                    dict(item) for item in cached_identity.payload.get("_source_catalog") or []
                    if isinstance(item, dict)
                )
                raw["identity"] = {
                    key: value for key, value in cached_identity.payload.items()
                    if key not in {"_source_catalog", "_profile_metadata"}
                }
                cache_reused_sections.append({
                    "section": "identity", "snapshot_id": cached_identity.snapshot_id,
                })
                self._mark_rule_cache_reuse(
                    raw.get("source_acquisition") or {},
                    slots={"sse_catalog", "szse_catalog", "annual_report"},
                    snapshot_id=cached_identity.snapshot_id,
                )

        methodology_required = ("version", "source_url")
        if any(
            (raw.get("index_methodology", {}).get(key) or {}).get("status") != "available"
            or not (raw.get("index_methodology", {}).get(key) or {}).get("source_ids")
            for key in methodology_required
        ):
            cached_methodology = self._latest_section_with_fields(
                normalized, "index_methodology", methodology_required
            )
            if cached_methodology is not None:
                cached_source_catalog.extend(
                    dict(item) for item in cached_methodology.payload.get("_source_catalog") or []
                    if isinstance(item, dict)
                )
                raw["index_methodology"] = {
                    key: value for key, value in cached_methodology.payload.items()
                    if key not in {"_source_catalog", "_profile_metadata"}
                }
                cache_reused_sections.append({
                    "section": "index_methodology",
                    "snapshot_id": cached_methodology.snapshot_id,
                })
                self._mark_rule_cache_reuse(
                    raw.get("source_acquisition") or {},
                    slots={"methodology"},
                    snapshot_id=cached_methodology.snapshot_id,
                )

        cached_metrics = self.store.latest_snapshot_by_created_at(
            normalized, "product_metrics"
        )
        if cached_metrics is not None:
            for key, current in list((raw.get("product_metrics") or {}).items()):
                cached = cached_metrics.payload.get(key)
                if (
                    isinstance(current, dict)
                    and current.get("status") != "available"
                    and isinstance(cached, dict)
                    and cached.get("status") == "available"
                ):
                    raw["product_metrics"][key] = {
                        **cached,
                        "status": "stale",
                        "note": "本轮来源未返回新值，保留上一不可变快照供用户核对",
                        "cache_snapshot_id": cached_metrics.snapshot_id,
                    }

        _finalize_product_profile_state(raw)
        origin_id = stable_fingerprint("etfprofileorigin", [normalized, raw.get("retrieved_at")])
        archived_sources = self._archive_sources(normalized, list(raw.get("sources") or []), origin_id)
        archived_sources.extend(cached_source_catalog)
        sources_by_id = {str(item.get("source_id") or ""): item for item in archived_sources}

        def sources_for(section: dict[str, Any]) -> list[dict[str, Any]]:
            ids = {
                str(source_id)
                for field in section.values() if isinstance(field, dict)
                for source_id in field.get("source_ids") or []
            }
            return [sources_by_id[source_id] for source_id in sorted(ids) if source_id in sources_by_id]

        identity_snapshot = self._save_section(
            symbol=normalized, snapshot_type="identity",
            data_as_of=_section_as_of(raw["identity"], str(raw["data_as_of"])),
            payload=dict(raw["identity"]), sources=sources_for(raw["identity"]), freshness_days=180,
        )
        methodology_snapshot = self._save_section(
            symbol=normalized, snapshot_type="index_methodology",
            data_as_of=_section_as_of(raw["index_methodology"], str(raw["data_as_of"])),
            payload=dict(raw["index_methodology"]), sources=sources_for(raw["index_methodology"]),
            freshness_days=180,
        )
        share_history_snapshot = None
        peer_group_snapshot = None
        share_tracking = None
        if include_share_flows:
            try:
                tracked_code = str((raw["identity"].get("tracked_index_code") or {}).get("value") or "")
                tracked_name = str((raw["identity"].get("tracked_index_name") or {}).get("value") or "")
                previous_peer = self.store.latest_snapshot_by_created_at(
                    normalized, "peer_group"
                )
                share_tracking = self.share_provider.fetch(
                    normalized, tracked_index_code=tracked_code,
                    tracked_index_name=tracked_name, as_of=as_of,
                    previous_peer_group=previous_peer,
                )
                tracking_errors = list(share_tracking.get("errors") or [])
                if tracking_errors:
                    raw.setdefault("source_errors", []).extend(
                        {"source": "share_tracking", "error": str(item)[:240]}
                        for item in tracking_errors
                    )
                share_sources = self._archive_sources(
                    normalized, list(share_tracking.pop("sources", []) or []), origin_id,
                )
                archived_sources.extend(share_sources)
                sources_by_id.update({
                    str(item.get("source_id") or ""): item
                    for item in share_sources if item.get("source_id")
                })
                subject_member = next(
                    (item for item in share_tracking.get("members") or [] if item.get("symbol") == normalized),
                    {},
                )
                official_units = _number(subject_member.get("current_units"))
                official_price = _number(subject_member.get("current_price"))
                official_as_of = str(
                    subject_member.get("data_as_of")
                    or share_tracking.get("data_as_of")
                    or raw["data_as_of"]
                )
                official_source_ids = [
                    str(source_id) for source_id in subject_member.get("source_ids") or []
                    if str(source_id) in sources_by_id
                ]
                prior_units = _number(
                    (raw.get("product_metrics", {}).get("fund_units") or {}).get("value")
                )
                if official_units is not None and official_source_ids:
                    if prior_units is not None and prior_units != official_units:
                        raw.setdefault("conflicts", []).append({
                            "field": "fund_units",
                            "lower_priority_value": prior_units,
                            "official_exchange_value": official_units,
                            "resolution": "official_exchange_end_of_day_preferred",
                            "data_as_of": official_as_of,
                            "source_ids": official_source_ids,
                        })
                    raw["product_metrics"]["fund_units"] = _field(
                        official_units,
                        source_ids=official_source_ids,
                        data_as_of=official_as_of,
                        semantics="official_exchange_end_of_day_fund_units",
                        unit="fund_units",
                        note=(
                            "交易所日终份额；与盘中行情源份额不一致时以交易所披露为准"
                            if prior_units is not None and prior_units != official_units else None
                        ),
                    )
                    if official_price is not None:
                        raw["product_metrics"]["exchange_market_value"] = _field(
                            official_units * official_price,
                            source_ids=official_source_ids,
                            data_as_of=official_as_of,
                            semantics="market_price_times_official_exchange_end_of_day_fund_units",
                            unit="CNY",
                            formula="fund_units * current_price",
                            input_metrics=("fund_units", "current_price"),
                            calculation_version="etf-product-calc-v1",
                        )
                history_payload = {
                    "tracked_index_code": tracked_code,
                    "tracked_index_name": tracked_name,
                    "observations": list(subject_member.get("history") or []),
                    "current_units": subject_member.get("current_units"),
                    "delta_1d": subject_member.get("delta_1d"),
                    "delta_5d": subject_member.get("delta_5d"),
                    "delta_20d": subject_member.get("delta_20d"),
                    "estimated_net_flow_1d": subject_member.get("estimated_net_flow_1d"),
                    "estimated_net_flow_semantics": subject_member.get("estimated_net_flow_semantics"),
                    "_source_catalog": _stable_source_catalog(share_sources),
                }
                share_history_snapshot = build_etf_snapshot(
                    symbol=normalized, snapshot_type="share_history",
                    data_as_of=str(share_tracking.get("data_as_of") or raw["data_as_of"]),
                    payload=history_payload,
                    coverage_ratio=(1.0 if subject_member.get("delta_1d") is not None else 0.5),
                    source_ids=[str(item.get("source_id")) for item in share_sources if item.get("source_id")],
                    freshness_expires_at=(datetime.now(timezone.utc) + timedelta(days=2)).isoformat(),
                    minimum_coverage=0.0,
                )
                share_history_snapshot = self.store.save_snapshot(share_history_snapshot)[0]
                peer_payload = {
                    **{key: value for key, value in share_tracking.items() if key not in {"errors"}},
                    "_source_catalog": _stable_source_catalog(share_sources),
                }
                peer_group_snapshot = build_etf_snapshot(
                    symbol=normalized, snapshot_type="peer_group",
                    data_as_of=str(share_tracking.get("data_as_of") or raw["data_as_of"]),
                    payload=peer_payload,
                    coverage_ratio=float(share_tracking.get("unit_change_coverage_ratio") or 0.0),
                    source_ids=[str(item.get("source_id")) for item in share_sources if item.get("source_id")],
                    freshness_expires_at=(datetime.now(timezone.utc) + timedelta(days=2)).isoformat(),
                    minimum_coverage=0.0,
                )
                peer_group_snapshot = self.store.save_snapshot(peer_group_snapshot)[0]
            except Exception as exc:
                raw.setdefault("source_errors", []).append({"source": "share_tracking", "error": str(exc)[:240]})

        _finalize_product_profile_state(raw)
        profile_metadata = {
            "hard_gate_status": str(raw.get("hard_gate_status") or "failed_validation"),
            "quality_status": str(raw.get("quality_status") or "failed_validation"),
            "missing_hard_fields": list(raw.get("missing_hard_fields") or []),
            "missing_optional_fields": list(raw.get("missing_optional_fields") or []),
            "conflicts": list(raw.get("conflicts") or []),
            "refresh_errors": list(raw.get("source_errors") or []),
            "source_policy": {
                "registry_version": self.source_registry.schema_version,
                "rules": sorted(
                    [
                        *list((raw.get("source_acquisition") or {}).get("rules") or []),
                        *list((share_tracking or {}).get("source_acquisition", {}).get("rules") or []),
                    ],
                    key=lambda item: (
                        str(item.get("phase") or ""),
                        str(item.get("rule_id") or ""),
                    ),
                ),
            },
            "cache_reused_sections": cache_reused_sections,
        }
        metrics_payload = {
            **dict(raw["product_metrics"]),
            "_profile_metadata": profile_metadata,
        }
        metrics_snapshot = self._save_section(
            symbol=normalized,
            snapshot_type="product_metrics",
            data_as_of=_section_as_of(metrics_payload, str(raw["data_as_of"])),
            payload=metrics_payload,
            sources=sources_for(raw["product_metrics"]),
            freshness_days=2,
        )
        profile = self._compose(
            identity_snapshot, methodology_snapshot, metrics_snapshot,
            share_history_snapshot=share_history_snapshot,
            peer_group_snapshot=peer_group_snapshot,
            sources=archived_sources,
            hard_gate_status=str(raw.get("hard_gate_status") or "failed_validation"),
            quality_status=str(raw.get("quality_status") or "failed_validation"),
            missing_hard_fields=list(raw.get("missing_hard_fields") or []),
            missing_optional_fields=list(raw.get("missing_optional_fields") or []),
            conflicts=list(raw.get("conflicts") or []),
            refresh_errors=list(raw.get("source_errors") or []),
        )
        return profile

    def _compose(
        self,
        identity: ETFResearchSnapshot,
        methodology: ETFResearchSnapshot,
        metrics: ETFResearchSnapshot,
        *,
        share_history_snapshot: ETFResearchSnapshot | None = None,
        peer_group_snapshot: ETFResearchSnapshot | None = None,
        sources: list[dict[str, Any]] | None = None,
        hard_gate_status: str | None = None,
        quality_status: str | None = None,
        missing_hard_fields: list[str] | None = None,
        missing_optional_fields: list[str] | None = None,
        conflicts: list[dict[str, Any]] | None = None,
        refresh_errors: list[Any] | None = None,
    ) -> dict[str, Any]:
        def clean_payload(snapshot: ETFResearchSnapshot) -> dict[str, Any]:
            return {
                key: value for key, value in snapshot.payload.items()
                if key not in {"_source_catalog", "_profile_metadata"}
            }

        snapshot_ids = {
            "identity": identity.snapshot_id,
            "index_methodology": methodology.snapshot_id,
            "product_metrics": metrics.snapshot_id,
        }
        if share_history_snapshot:
            snapshot_ids["share_history"] = share_history_snapshot.snapshot_id
        if peer_group_snapshot:
            snapshot_ids["peer_group"] = peer_group_snapshot.snapshot_id
        identity_payload = clean_payload(identity)
        methodology_payload = clean_payload(methodology)
        metrics_payload = clean_payload(metrics)
        restored_sources = [
            {**dict(item), "retrieved_at": snapshot.retrieved_at}
            for snapshot in (
                identity, methodology, metrics, share_history_snapshot, peer_group_snapshot,
            )
            if snapshot is not None
            for item in snapshot.payload.get("_source_catalog") or []
            if isinstance(item, dict)
        ]
        share_payload = clean_payload(share_history_snapshot) if share_history_snapshot else None
        peer_payload = clean_payload(peer_group_snapshot) if peer_group_snapshot else None
        persisted_metadata = dict(metrics.payload.get("_profile_metadata") or {})
        resolved_hard_gate_status = (
            hard_gate_status
            if hard_gate_status is not None
            else persisted_metadata.get("hard_gate_status")
        )
        resolved_quality_status = (
            quality_status
            if quality_status is not None
            else persisted_metadata.get("quality_status")
        )
        resolved_missing_hard_fields = list(
            missing_hard_fields
            if missing_hard_fields is not None
            else persisted_metadata.get("missing_hard_fields") or []
        )
        resolved_missing_optional_fields = list(
            missing_optional_fields
            if missing_optional_fields is not None
            else persisted_metadata.get("missing_optional_fields") or []
        )
        resolved_conflicts = list(
            conflicts
            if conflicts is not None
            else persisted_metadata.get("conflicts") or []
        )
        resolved_refresh_errors = list(
            refresh_errors
            if refresh_errors is not None
            else persisted_metadata.get("refresh_errors") or []
        )
        resolved_source_policy = dict(persisted_metadata.get("source_policy") or {})
        resolved_cache_reused_sections = list(
            persisted_metadata.get("cache_reused_sections") or []
        )
        semantic = {
            "symbol": identity.symbol, "snapshot_ids": snapshot_ids,
            "identity": identity_payload, "index_methodology": methodology_payload,
            "product_metrics": metrics_payload,
            "share_history": share_payload,
            "peer_group": peer_payload,
            "quality_metadata": {
                "hard_gate_status": resolved_hard_gate_status,
                "quality_status": resolved_quality_status,
                "missing_hard_fields": resolved_missing_hard_fields,
                "missing_optional_fields": resolved_missing_optional_fields,
                "conflicts": resolved_conflicts,
                "refresh_errors": resolved_refresh_errors,
                "source_policy": resolved_source_policy,
                "cache_reused_sections": resolved_cache_reused_sections,
            },
        }
        profile_id = stable_fingerprint("etfprofile", semantic)
        errors = resolved_refresh_errors
        return {
            "schema_version": 1, "profile_snapshot_id": profile_id,
            "symbol": identity.symbol,
            "data_as_of": max(identity.data_as_of, methodology.data_as_of, metrics.data_as_of),
            "retrieved_at": max(identity.retrieved_at, methodology.retrieved_at, metrics.retrieved_at),
            "snapshot_ids": snapshot_ids,
            "identity": identity_payload, "index_methodology": methodology_payload,
            "product_metrics": metrics_payload,
            "share_history": share_payload,
            "peer_group": peer_payload,
            "sources": list({
                str(item.get("source_id")): item
                for item in [*restored_sources, *(sources or [])]
                if item.get("source_id")
            }.values()),
            "hard_gate_status": resolved_hard_gate_status or (
                "passed" if identity.quality_status != "failed_validation" and methodology.quality_status != "failed_validation"
                else "failed_validation"
            ),
            "quality_status": resolved_quality_status or (
                "passed" if metrics.quality_status == "passed" else "passed_with_gaps"
            ),
            "missing_hard_fields": resolved_missing_hard_fields,
            "missing_optional_fields": resolved_missing_optional_fields,
            "conflicts": resolved_conflicts,
            "refresh_errors": errors,
            "refresh_status": "completed_with_gaps" if errors else "completed",
            "source_policy": resolved_source_policy,
            "cache_reused_sections": resolved_cache_reused_sections,
        }

    def source_plan(
        self,
        symbol: str,
        *,
        manager: str | None = None,
        index_code: str | None = None,
        as_of: str | None = None,
    ) -> dict[str, Any]:
        """Expose the reusable rules selected for a symbol without fetching them."""

        context = {
            **source_context(symbol, manager=manager, index_code=index_code),
            "data_as_of": _as_of_date(as_of).isoformat(),
        }
        return {
            "registry_version": self.source_registry.schema_version,
            "symbol": normalize_etf_symbol(symbol),
            "phases": {
                phase: self.source_registry.plan(phase=phase, context=context)
                for phase in ("product_profile", "share_flow")
            },
        }

    def latest_profile(self, symbol: str) -> dict[str, Any] | None:
        normalized = normalize_etf_symbol(symbol)
        identity = self.store.latest_snapshot_by_created_at(normalized, "identity")
        methodology = self.store.latest_snapshot_by_created_at(
            normalized, "index_methodology"
        )
        metrics = self.store.latest_snapshot_by_created_at(
            normalized, "product_metrics"
        )
        if not identity or not methodology or not metrics:
            return None
        return self._compose(
            identity, methodology, metrics,
            share_history_snapshot=self.store.latest_snapshot_by_created_at(
                normalized, "share_history"
            ),
            peer_group_snapshot=self.store.latest_snapshot_by_created_at(
                normalized, "peer_group"
            ),
        )

    def get_or_refresh(
        self,
        symbol: str,
        *,
        force_refresh: bool = False,
        as_of: str | None = None,
        instrument_profile: dict[str, Any] | None = None,
        universe_snapshot: ETFResearchSnapshot | None = None,
    ) -> dict[str, Any]:
        latest = self.latest_profile(symbol)
        requested_day = str(as_of or datetime.now(timezone.utc).date().isoformat())[:10]
        metric_days = sorted(
            str(item.get("data_as_of") or "")[:10]
            for item in dict((latest or {}).get("product_metrics") or {}).values()
            if isinstance(item, dict)
            and item.get("status") == "available"
            and re.fullmatch(r"20\d{2}-\d{2}-\d{2}", str(item.get("data_as_of") or "")[:10])
        )
        metrics_fresh = False
        if metric_days:
            try:
                metrics_fresh = (
                    date.fromisoformat(requested_day) - date.fromisoformat(metric_days[-1])
                ).days <= 2
            except ValueError:
                metrics_fresh = False
        if latest is not None and not force_refresh and metrics_fresh:
            return {
                **latest,
                "cache_hit": True,
                "refresh_status": "cache_only",
                "refresh_plan": {
                    "policy": "field_freshness_v1",
                    "identity": "version_reuse",
                    "product_metrics": "fresh_cache_reuse",
                    "requested_as_of": requested_day,
                },
            }
        try:
            return {
                **self.refresh(
                    symbol, as_of=as_of, instrument_profile=instrument_profile,
                    universe_snapshot=universe_snapshot,
                ),
                "cache_hit": False,
                "refresh_plan": {
                    "policy": "field_freshness_v1",
                    "identity": "version_or_missing",
                    "product_metrics": "refresh_due_to_force_or_staleness",
                    "requested_as_of": requested_day,
                },
            }
        except Exception as exc:
            if latest is None:
                raise
            return {
                **latest, "cache_hit": True, "refresh_status": "completed_with_gaps",
                "refresh_errors": [*list(latest.get("refresh_errors") or []), str(exc)[:240]],
                "stale": True,
            }

    @staticmethod
    def to_report_records(
        profile: dict[str, Any],
        *,
        base_facts: Iterable[dict[str, Any]] = (),
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        symbol = str(profile.get("symbol") or "").upper()
        evidence: list[dict[str, Any]] = []
        for source in profile.get("sources") or []:
            if not isinstance(source, dict) or not source.get("source_id"):
                continue
            evidence.append({
                "evidence_id": str(source["source_id"]), "symbol": symbol,
                "domain": str(source.get("kind") or "etf_product"),
                "source": str(source.get("publisher") or "官方来源"),
                "source_locator": str(source.get("url") or source.get("document_ref") or ""),
                "retrieved_at": str(source.get("retrieved_at") or profile.get("retrieved_at") or utc_now()),
                "published_at": source.get("published_at"),
                "content_hash": str(source.get("content_hash") or ""),
                "summary": str(source.get("title") or "ETF 产品资料"),
                "status": (
                    "verified" if source.get("verification_status") == "official_primary" else "partial"
                ),
                "metadata": {
                    "verification_status": source.get("verification_status"),
                    "body_status": source.get("body_status"),
                    "document_ref": source.get("document_ref"),
                    "source_rule_id": dict(source.get("metadata") or {}).get("source_rule_id"),
                    "parser_id": dict(source.get("metadata") or {}).get("parser_id"),
                    "registry_version": dict(source.get("metadata") or {}).get("registry_version"),
                },
            })
        evidence_ids = {str(item["evidence_id"]) for item in evidence}
        facts: list[dict[str, Any]] = []
        pending_facts: list[dict[str, Any]] = []
        for section_name in ("identity", "index_methodology", "product_metrics"):
            section = dict(profile.get(section_name) or {})
            for key, raw in section.items():
                if not isinstance(raw, dict) or raw.get("status") != "available":
                    continue
                value = raw.get("value")
                linked = [str(item) for item in raw.get("source_ids") or [] if str(item) in evidence_ids]
                provisional_fact_id = stable_fingerprint(
                    "fact", [symbol, section_name, key, value, raw.get("data_as_of"), linked]
                )
                pending_facts.append({
                    "fact_id": provisional_fact_id, "symbol": symbol, "metric": key,
                    "value": str(value), "unit": str(raw.get("unit") or "text"),
                    "period": str(raw.get("data_as_of") or profile.get("data_as_of") or ""),
                    "formula": raw.get("formula"),
                    "input_metric_keys": list(raw.get("input_metrics") or []),
                    "input_fact_ids": [],
                    "evidence_ids": linked,
                    "calculation_version": (
                        raw.get("calculation_version") or "etf-product-profile-v2"
                    ),
                    "validation_status": "pass" if linked else "warning",
                    "metadata": {
                        "section": section_name, "label": _FIELD_LABELS.get(key, key),
                        "semantics": raw.get("semantics"), "note": raw.get("note"),
                        "source_kind": raw.get("source_kind") or (
                            "derived" if raw.get("formula") else "observed"
                        ),
                    },
                })

        facts_by_metric = {
            str(item.get("metric") or ""): dict(item)
            for item in [*list(base_facts), *pending_facts]
            if isinstance(item, dict) and item.get("metric") and item.get("fact_id")
        }
        for item in pending_facts:
            input_keys = list(item.pop("input_metric_keys", []) or [])
            inputs = [
                str(facts_by_metric[key]["fact_id"])
                for key in input_keys
                if key in facts_by_metric
            ]
            item["input_fact_ids"] = inputs
            if item.get("formula"):
                item["fact_id"] = stable_fingerprint(
                    "fact",
                    [
                        symbol,
                        item.get("metric"),
                        item.get("value"),
                        item.get("period"),
                        item.get("formula"),
                        inputs,
                    ],
                )
                if len(inputs) != len(input_keys):
                    item["validation_status"] = "warning"
                    item["metadata"]["lineage_status"] = "incomplete"
                    item["metadata"]["missing_input_metrics"] = [
                        key for key in input_keys if key not in facts_by_metric
                    ]
                else:
                    item["metadata"]["lineage_status"] = "replayable"
                    linked = set(item.get("evidence_ids") or [])
                    for key in input_keys:
                        linked.update(facts_by_metric[key].get("evidence_ids") or [])
                    item["evidence_ids"] = sorted(linked)
            facts.append(item)

        share = dict(profile.get("share_history") or {})
        peer = dict(profile.get("peer_group") or {})

        def append_flow_fact(
            *,
            fact_symbol: str,
            metric: str,
            value: Any,
            unit: str,
            period: str,
            linked: list[str],
            semantics: str,
            formula: str | None = None,
            input_fact_ids: list[str] | None = None,
            metadata: dict[str, Any] | None = None,
        ) -> str | None:
            if value is None:
                return None
            inputs = list(input_fact_ids or [])
            fact_id = stable_fingerprint(
                "fact",
                [fact_symbol, metric, value, period, linked, formula, inputs],
            )
            facts.append({
                "fact_id": fact_id,
                "symbol": fact_symbol,
                "metric": metric,
                "value": str(value),
                "unit": unit,
                "period": period,
                "formula": formula,
                "input_fact_ids": inputs,
                "evidence_ids": linked,
                "calculation_version": "etf-peer-flow-v1",
                "validation_status": "pass" if linked else "warning",
                "metadata": {
                    "section": "share_tracking",
                    "semantics": semantics,
                    **dict(metadata or {}),
                },
            })
            return fact_id

        member_lineage: dict[str, dict[str, str | None]] = {}
        for member in peer.get("members") or []:
            if not isinstance(member, dict):
                continue
            member_symbol = str(member.get("symbol") or "").upper()
            if not member_symbol:
                continue
            linked = sorted({
                str(source_id) for source_id in member.get("source_ids") or []
                if str(source_id) in evidence_ids
            })
            period = str(member.get("data_as_of") or peer.get("data_as_of") or "")
            metadata = {
                "peer_symbol": member_symbol,
                "peer_name": member.get("name"),
                "mapping_status": member.get("mapping_status"),
                "estimation_price_type": member.get("estimation_price_type"),
            }
            history = [
                item for item in member.get("history") or [] if isinstance(item, dict)
            ]
            current_id = append_flow_fact(
                fact_symbol=member_symbol,
                metric="peer_member_current_fund_units",
                value=member.get("current_units"),
                unit="fund_units",
                period=period,
                linked=linked,
                semantics="official_exchange_end_of_day_fund_units",
                metadata=metadata,
            )
            prior = history[1] if len(history) > 1 else {}
            prior_linked = sorted({
                str(source_id) for source_id in prior.get("source_ids") or []
                if str(source_id) in evidence_ids
            }) or linked
            prior_id = append_flow_fact(
                fact_symbol=member_symbol,
                metric="peer_member_prior_fund_units",
                value=prior.get("fund_units"),
                unit="fund_units",
                period=str(prior.get("data_as_of") or period),
                linked=prior_linked,
                semantics="official_exchange_prior_trading_day_fund_units",
                metadata=metadata,
            )
            delta_inputs = [item for item in (current_id, prior_id) if item]
            delta_id = append_flow_fact(
                fact_symbol=member_symbol,
                metric="peer_member_fund_units_change_1d",
                value=member.get("delta_1d"),
                unit="fund_units",
                period=period,
                linked=sorted(set(linked + prior_linked)),
                semantics="current_units_minus_prior_trading_day_units",
                formula=(
                    "current_units - prior_trading_day_units"
                    if len(delta_inputs) == 2 else None
                ),
                input_fact_ids=delta_inputs if len(delta_inputs) == 2 else [],
                metadata=metadata,
            )
            estimation_price_type = str(member.get("estimation_price_type") or "")
            estimation_price = member.get("estimation_price")
            if estimation_price is None:
                estimation_price = member.get("current_price")
            price_id = append_flow_fact(
                fact_symbol=member_symbol,
                metric=(
                    "peer_member_nav_proxy"
                    if estimation_price_type == "exchange_published_nav_proxy"
                    else "peer_member_market_price"
                ),
                value=estimation_price,
                unit="CNY_per_fund_unit",
                period=period,
                linked=linked,
                semantics=(
                    "official_exchange_published_unit_nav_proxy"
                    if estimation_price_type == "exchange_published_nav_proxy"
                    else "official_exchange_current_market_price"
                ),
                metadata=metadata,
            )
            flow_inputs = [item for item in (delta_id, price_id) if item]
            flow_id = append_flow_fact(
                fact_symbol=member_symbol,
                metric="peer_member_estimated_net_flow_1d",
                value=member.get("estimated_net_flow_1d"),
                unit="CNY",
                period=period,
                linked=linked,
                semantics=str(member.get("estimated_net_flow_semantics") or ""),
                formula=("share_delta_1d * estimation_price" if len(flow_inputs) == 2 else None),
                input_fact_ids=flow_inputs if len(flow_inputs) == 2 else [],
                metadata=metadata,
            )
            member_lineage[member_symbol] = {
                "current": current_id,
                "delta": delta_id,
                "price": price_id,
                "flow": flow_id,
            }

        subject_lineage = member_lineage.get(symbol) or {}
        all_linked = sorted({
            evidence_id
            for member in peer.get("members") or [] if isinstance(member, dict)
            for evidence_id in member.get("source_ids") or []
            if str(evidence_id) in evidence_ids
        })
        peer_period = str(peer.get("data_as_of") or profile.get("data_as_of") or "")
        for key, value, unit, semantics, input_id in (
            ("etf_fund_units", share.get("current_units"), "fund_units", "official_exchange_fund_units", subject_lineage.get("current")),
            ("etf_fund_units_change_1d", share.get("delta_1d"), "fund_units", "current_units_minus_prior_trading_day_units", subject_lineage.get("delta")),
            ("etf_estimated_net_flow_1d", share.get("estimated_net_flow_1d"), "CNY", str(share.get("estimated_net_flow_semantics") or ""), subject_lineage.get("flow")),
        ):
            inputs = [str(input_id)] if input_id else []
            append_flow_fact(
                fact_symbol=symbol,
                metric=key,
                value=value,
                unit=unit,
                period=peer_period,
                linked=all_linked,
                semantics=semantics,
                formula="same_as_subject_peer_observation" if inputs else None,
                input_fact_ids=inputs,
            )

        flow_inputs = [
            str(item.get("flow")) for item in member_lineage.values() if item.get("flow")
        ]
        delta_inputs = [
            str(item.get("delta")) for item in member_lineage.values() if item.get("delta")
        ]
        current_inputs = [
            str(item.get("current")) for item in member_lineage.values() if item.get("current")
        ]
        for key, value, unit, semantics, formula, inputs in (
            ("peer_group_estimated_net_flow_1d", peer.get("estimated_net_flow_1d"), "CNY", str(peer.get("estimated_net_flow_semantics") or ""), "sum(peer_member_estimated_net_flow_1d)", flow_inputs),
            ("peer_group_inflow_member_ratio_1d", peer.get("inflow_member_ratio_1d"), "ratio", "positive_share_delta_members_divided_by_comparable_members", "positive_delta_members / comparable_members", delta_inputs),
            ("peer_group_member_count", peer.get("member_count"), "count", "same_tracked_index_peer_count", None, []),
            ("peer_group_unit_change_coverage", peer.get("unit_change_coverage_ratio"), "ratio", "comparable_share_change_members_divided_by_group_members", "comparable_members / group_members", current_inputs),
        ):
            append_flow_fact(
                fact_symbol=symbol,
                metric=key,
                value=value,
                unit=unit,
                period=peer_period,
                linked=all_linked,
                semantics=semantics,
                formula=formula if inputs else None,
                input_fact_ids=inputs,
            )
        return facts, evidence


_shared_service: ETFProductProfileService | None = None
_shared_lock = threading.Lock()


def get_etf_product_profile_service() -> ETFProductProfileService:
    global _shared_service
    with _shared_lock:
        if _shared_service is None:
            _shared_service = ETFProductProfileService()
        return _shared_service
